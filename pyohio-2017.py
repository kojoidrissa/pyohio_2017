
# coding: utf-8

# In[ ]:


# %load simple_snakesheets.py
# This is some of my basic demo code from the 2017-04-18 meetup

import openpyxl
from openpyxl import Workbook

# get workbook object; the data_only option captures the current value of any formulae
# xyz.xlsx needs to be replaced with the name of the file you're trying to read. That file should be in the same directory as this code, OR you need to provide the FULL path to that file, not just it's name
wb = openpyxl.load_workbook('xyz.xlsx', data_only= True)

# This assigns the data from a specific sheet to variable that you can work with. Replace sheet1 inside the "get_sheet_by_name" function with the name of a spreadSHEET...an individual tab.
sheet1 = wb.get_sheet_by_name('sheet1')
sheet2 = wb.get_sheet_by_name('sheet2')
sheet3 = wb.get_sheet_by_name('sheet3')

# this will print the value of cell B2 in Sheet 1. Make sure there's a value there.
print(sheet1["B2"].value)

# This is how you create a new blank workbook, create a worksheet inside that workbook with a specific name (Kojo), set a value in a specific cell in that worksheet, then save the workbook to the filename you choose. 
wb_out = Workbook()
ws1 = wb_out.create_sheet("Kojo")
ws1["b2"] = 42
wb_out.save('kojo_demo.xlsx')


# In[ ]:





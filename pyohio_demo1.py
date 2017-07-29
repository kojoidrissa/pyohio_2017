
# coding: utf-8

# In[ ]:


# %load pyohio_demo1.py
# This is Demo 1 for my the 2017-07-29 PyOhio talk

import openpyxl
from openpyxl import Workbook

# get workbook object; the data_only option captures the current value of any formulae
# Your file should be in the same directory as this code, OR you need to provide the FULL path to that file, not just it's name
wb = openpyxl.load_workbook('pyxl_demo_pyohio.xlxs', data_only= True)

# Get the names of the sheets (tabs) in the workbook
print (wb.sheetnames)

# What can you do with a workbook?
dir(wb)

demo_worksheet = wb.get_sheet_by_name("clean_data")
# If you've added a new worksheet since you loaded the workbook, you'll have to reload it.

# List of employee IDs
employee1 = {row[0].value for row in demo_worksheet.rows if row[0].value != 'employee_num'}

# This assigns the data from a specific sheet to variable that you can work with. Replace sheet1 inside the "get_sheet_by_name" function with the name of a spreadSHEET...an individual tab.
sheet1 = wb.get_sheet_by_name('sheet1')
sheet2 = wb.get_sheet_by_name('sheet2')
sheet3 = wb.get_sheet_by_name('sheet3')

# this will print the value of cell B2 in Sheet 1. Make sure there's a value there.
print(sheet1["B2"].value)

# This is how you create a new blank workbook, create a worksheet inside that workbook with a specific name (Kojo), set a value in a specific cell in that worksheet, then save the workbook to the filename you choose. 
wb_out = Workbook()
ws1 = wb_out.create_sheet("Kojo")
ws1["b2"] = 42
wb_out.save('kojo_demo.xlsx')


# In[ ]:




